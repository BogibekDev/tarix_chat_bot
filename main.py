import asyncio
import os
import io
from datetime import datetime
from typing import Optional, Iterable
import openpyxl
from openpyxl.utils import get_column_letter


import asyncpg
from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart, Command
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    Message, User as TgUser, CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup, FSInputFile
)
from dotenv import load_dotenv

# =========================
# Config
# =========================
load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")
if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN yo'q. .env to'ldir.")

DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL yo'q. .env to'ldir.")

# Seed ro'yxatlari (faqat boshlang'ich bosqichda DBga yozamiz)
def _parse_id_list(s: str | None) -> list[int]:
    if not s:
        return []
    out = []
    for x in s.split(","):
        x = x.strip()
        if x.isdigit():
            out.append(int(x))
    return out

SEED_SUPERADMINS = _parse_id_list(os.getenv("SUPERADMIN_IDS"))
SEED_ADMINS = _parse_id_list(os.getenv("ADMIN_IDS"))

bot = Bot(BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())

# Eksport fayl nomi (diskga yozmaymiz, lekin nom zarur)

# =========================
# DB Layer
# =========================
class DB:
    def __init__(self, dsn: str):
        self._dsn = dsn
        self.pool: asyncpg.Pool | None = None

    async def connect(self):
        self.pool = await asyncpg.create_pool(self._dsn, min_size=1, max_size=10)

    async def close(self):
        if self.pool:
            await self.pool.close()

    async def ensure_schema(self):
        sql = """
        CREATE TABLE IF NOT EXISTS users (
          tg_id BIGINT PRIMARY KEY,
          username TEXT,
          first_name TEXT,
          last_name TEXT,
          created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS admins (
          tg_id BIGINT PRIMARY KEY,
          role TEXT NOT NULL CHECK (role IN ('admin','super')),
          created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS idx_users_username ON users ((lower(coalesce(username,''))));
        """
        async with self.pool.acquire() as con:
            await con.execute(sql)

    async def seed_admins(self, super_ids: Iterable[int], admin_ids: Iterable[int]):
        # insert if not exists
        async with self.pool.acquire() as con:
            async with con.transaction():
                for uid in set(super_ids or []):
                    await con.execute(
                        "INSERT INTO admins(tg_id, role) VALUES($1,'super') ON CONFLICT (tg_id) DO NOTHING",
                        uid,
                    )
                for uid in set(admin_ids or []):
                    # agar super bo'lsa, yana admin qo'ymaymiz
                    await con.execute(
                        """
                        INSERT INTO admins(tg_id, role) 
                        SELECT $1,'admin'
                        WHERE NOT EXISTS (SELECT 1 FROM admins WHERE tg_id=$1)
                        """,
                        uid,
                    )

    # --- Users ---
    async def upsert_user(self, u: TgUser):
        async with self.pool.acquire() as con:
            await con.execute(
                """
                INSERT INTO users(tg_id, username, first_name, last_name)
                VALUES($1, $2, $3, $4)
                ON CONFLICT (tg_id) DO UPDATE
                  SET username=EXCLUDED.username,
                      first_name=EXCLUDED.first_name,
                      last_name=EXCLUDED.last_name
                """,
                u.id, u.username, u.first_name, u.last_name
            )

    async def list_user_ids(self) -> list[int]:
        async with self.pool.acquire() as con:
            rows = await con.fetch("SELECT tg_id FROM users ORDER BY tg_id")
            return [r["tg_id"] for r in rows]

    async def find_user_id_by_username(self, token: str) -> Optional[int]:
        # token: '@username' yoki 'username' bo'lishi mumkin
        token = token.strip()
        if token.startswith("@"):
            token = token[1:]
        token = token.lower()
        if not token:
            return None
        async with self.pool.acquire() as con:
            row = await con.fetchrow(
                "SELECT tg_id FROM users WHERE lower(coalesce(username,''))=$1",
                token
            )
            return row["tg_id"] if row else None

    async def export_users_xlsx_bytes(self) -> bytes:
        async with self.pool.acquire() as con:
            rows = await con.fetch(
                "SELECT tg_id, username, first_name, last_name, created_at FROM users ORDER BY tg_id"
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"

        # Sarlavhalar
        headers = ["Telegram ID", "Username", "First name", "Last name", "Created at"]
        ws.append(headers)

        # Ma’lumotlar
        for r in rows:
            ws.append([
                r["tg_id"],
                f"@{r['username']}" if r["username"] else "-",
                r["first_name"] or "-",
                r["last_name"] or "-",
                r["created_at"].strftime("%Y-%m-%d %H:%M:%S"),
            ])

        # Avtomatik ustun kengligi
        for i, col in enumerate(ws.columns, 1):
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 2

        # RAMga yozamiz
        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()


    # --- Admins ---
    async def list_admin_ids(self) -> set[int]:
        async with self.pool.acquire() as con:
            rows = await con.fetch("SELECT tg_id FROM admins WHERE role IN ('admin','super')")
            return {r["tg_id"] for r in rows}

    async def list_super_ids(self) -> set[int]:
        async with self.pool.acquire() as con:
            rows = await con.fetch("SELECT tg_id FROM admins WHERE role='super'")
            return {r["tg_id"] for r in rows}

    async def is_admin(self, uid: int) -> bool:
        async with self.pool.acquire() as con:
            row = await con.fetchrow("SELECT 1 FROM admins WHERE tg_id=$1", uid)
            return bool(row)

    async def is_super(self, uid: int) -> bool:
        async with self.pool.acquire() as con:
            row = await con.fetchrow("SELECT 1 FROM admins WHERE tg_id=$1 AND role='super'", uid)
            return bool(row)

    async def add_admin(self, uid: int) -> bool:
        # superga aylantirmaymiz, oddiy admin
        async with self.pool.acquire() as con:
            try:
                await con.execute(
                    "INSERT INTO admins(tg_id, role) VALUES($1,'admin')",
                    uid
                )
                return True
            except asyncpg.UniqueViolationError:
                return False

    async def remove_admin(self, uid: int) -> bool:
        # superadminni o'chirmaymiz (xavfsizlik)
        async with self.pool.acquire() as con:
            async with con.transaction():
                row = await con.fetchrow("SELECT role FROM admins WHERE tg_id=$1", uid)
                if not row:
                    return False
                if row["role"] == "super":
                    return False
                await con.execute("DELETE FROM admins WHERE tg_id=$1", uid)
                return True

    async def list_admins_text(self) -> str:
        async with self.pool.acquire() as con:
            supers = await con.fetch(
                "SELECT tg_id FROM admins WHERE role='super' ORDER BY tg_id"
            )
            admins = await con.fetch(
                "SELECT tg_id FROM admins WHERE role='admin' ORDER BY tg_id"
            )
        s_supers = "\n".join([f"• <code>{r['tg_id']}</code>" for r in supers]) or "—"
        s_admins = "\n".join([f"• <code>{r['tg_id']}</code>" for r in admins]) or "—"
        return f"<b>👑 Superadmins</b>\n{s_supers}\n\n<b>🛡 Admins</b>\n{s_admins}"


db = DB(DATABASE_URL)

# =========================
# States
# =========================
class ReplyStates(StatesGroup):
    waiting_text = State()

class BroadcastStates(StatesGroup):
    waiting_content = State()
    confirm_content = State()   # ✅ yangi state


class AdminMgmtStates(StatesGroup):
    waiting_add = State()
    waiting_remove = State()



@dp.callback_query(F.data == "panel:broadcast")
async def panel_broadcast(cq: CallbackQuery, state: FSMContext):
    if not await is_admin(cq.from_user.id):
        return
    await state.set_state(BroadcastStates.waiting_content)
    await cq.message.answer(
        "Broadcast rejimi: Menga <b>bitta xabar</b> yubor.\n"
        "Yuborilishini tasdiqlashdan oldin xabarni ko‘ramiz.\n"
    )
    await cq.answer()



@dp.message(BroadcastStates.waiting_content)
async def broadcast_preview(msg: Message, state: FSMContext):
    if not await is_admin(msg.from_user.id):
        return

    # Xabarni keyin yuborish uchun saqlab qo'yamiz
    await state.update_data(draft_msg=msg)

    # Adminni tasdiqlashga chaqiramiz
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ Ha, yubor!", callback_data="broadcast:confirm"),
            InlineKeyboardButton(text="❌ Yo‘q, bekor", callback_data="broadcast:cancel")
        ]
    ])
    await msg.answer("Shu xabarni barcha foydalanuvchilarga yuborishni hohlaysizmi?", reply_markup=kb)

    await state.set_state(BroadcastStates.confirm_content)

@dp.callback_query(F.data == "broadcast:confirm")
async def broadcast_confirm(cq: CallbackQuery, state: FSMContext):
    if not await is_admin(cq.from_user.id):
        return

    data = await state.get_data()
    draft: Message = data.get("draft_msg")
    if not draft:
        await cq.message.answer("❌ Xabar topilmadi, qaytadan urin.")
        await state.clear()
        return

    user_ids = await db.list_user_ids()
    total = len(user_ids)
    ok = 0
    fail = 0

    sem = asyncio.Semaphore(20)

    async def send_copy(uid: int):
        nonlocal ok, fail
        async with sem:
            try:
                await draft.copy_to(chat_id=uid)
                ok += 1
            except Exception:
                fail += 1

    await asyncio.gather(*(send_copy(uid) for uid in user_ids))

    await state.clear()
    await cq.message.answer(
        f"✅ Broadcast yakunlandi.\nYuborildi: <b>{ok}</b> / Jami: {total} / Xato: {fail}"
    )
    await cq.answer()

@dp.callback_query(F.data == "broadcast:cancel")
async def broadcast_cancel(cq: CallbackQuery, state: FSMContext):
    if not await is_admin(cq.from_user.id):
        return
    await state.clear()
    await cq.message.answer("❌ Broadcast bekor qilindi.")
    await cq.answer()

# =========================
# Keyboards
# =========================
def panel_kb(superadmin: bool):
    rows = [
        [InlineKeyboardButton(text="✉️ Broadcast", callback_data="panel:broadcast")],
        [InlineKeyboardButton(text="👥 Users count", callback_data="panel:count")],
        [InlineKeyboardButton(text="📤 Export users.xlsx", callback_data="panel:export")],
    ]
    if superadmin:
        rows.append([InlineKeyboardButton(text="👑 Manage Admins", callback_data="panel:admins")])
    return InlineKeyboardMarkup(inline_keyboard=rows)

def reply_kb(user_id: int, superadmin: bool):
    row = [InlineKeyboardButton(text="✍️ Javob qaytarish", callback_data=f"reply:{user_id}")]
    rows = [row]
    if superadmin:
        rows.append([
            InlineKeyboardButton(text="⚙️ Admin qilish", callback_data=f"admins:promote:{user_id}"),
            InlineKeyboardButton(text="🗑️ Adminlikdan o'chirish", callback_data=f"admins:revoke:{user_id}"),
        ])
    return InlineKeyboardMarkup(inline_keyboard=rows)

def admins_menu_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ Add admin", callback_data="admins:add")],
        [InlineKeyboardButton(text="➖ Remove admin", callback_data="admins:remove")],
        [InlineKeyboardButton(text="📄 List admins", callback_data="admins:list")],
        [InlineKeyboardButton(text="↩️ Back", callback_data="panel:back")],
    ])

# =========================
# Role utils (DB-based)
# =========================
async def all_admin_ids() -> set[int]:
    return await db.list_admin_ids()

async def is_superadmin(user_id: int) -> bool:
    return await db.is_super(user_id)

async def is_admin(user_id: int) -> bool:
    return await db.is_admin(user_id)

# =========================
# Handlers
# =========================
@dp.message(Command("admin"))
async def cmd_panel(msg: Message):
    if not await is_admin(msg.from_user.id):
        return
    await msg.answer(
        "Admin panel:",
        reply_markup=panel_kb(superadmin=await is_superadmin(msg.from_user.id)),
    )

@dp.callback_query(F.data == "panel:back")
async def panel_back(cq: CallbackQuery):
    if not await is_admin(cq.from_user.id):
        return
    await cq.message.edit_text(
        "Admin panel:",
        reply_markup=panel_kb(superadmin=await is_superadmin(cq.from_user.id)),
    )
    await cq.answer()

@dp.callback_query(F.data.startswith("panel:"))
async def panel_actions(cq: CallbackQuery, state: FSMContext):
    if not await is_admin(cq.from_user.id):
        return
    action = cq.data.split(":", 1)[1]

    if action == "broadcast":
        await state.set_state(BroadcastStates.waiting_content)
        await cq.message.answer(
            "📢 Broadcast rejimi boshlandi.\n"
            "Menga bir dona xabar yuboring, men uni barcha foydalanuvchilarga tarqataman.\n\n"
            "❌ Agar fikringizdan qaytsangiz, /cancel buyrug‘ini yuboring."
        )
    elif action == "export":
        data = await db.export_users_xlsx_bytes()
        tmp_path = "users.xlsx"
        with open(tmp_path, "wb") as f:
            f.write(data)
        try:
            doc = FSInputFile(tmp_path)
            await cq.message.answer_document(document=doc, caption="users.xlsx")
        finally:
            try:
                os.remove(tmp_path)
            except Exception:
                pass

    elif action == "admins":
        if not await is_superadmin(cq.from_user.id):
            await cq.answer("Faqat superadmin uchun.", show_alert=True)
            return
        await cq.message.answer("👑 Admin Management", reply_markup=admins_menu_kb())
    await cq.answer()

@dp.message(Command("cancel"))
async def cancel_any(msg: Message, state: FSMContext):
    if not await is_admin(msg.from_user.id):
        return
    await state.clear()
    await msg.answer("Bekor qilindi. Normal rejimga qaytdik.")

# ---------- Broadcast (parallel, throttled) ----------
@dp.message(BroadcastStates.waiting_content)
async def do_broadcast(msg: Message, state: FSMContext):
    if not await is_admin(msg.from_user.id):
        return

    user_ids = await db.list_user_ids()
    total = len(user_ids)

    # Tezkor lekin ehtiyot: 20 ta parallel oqim
    sem = asyncio.Semaphore(20)
    ok = 0
    fail = 0

    async def send_copy(uid: int):
        nonlocal ok, fail
        async with sem:
            try:
                await msg.copy_to(chat_id=uid)
                ok += 1
            except Exception:
                fail += 1

    await asyncio.gather(*(send_copy(uid) for uid in user_ids))

    await state.clear()
    await msg.answer(f"Broadcast yakunlandi.\nYuborildi: <b>{ok}</b> / Jami: {total} / Xato: {fail}")

# ---------- Reply flow ----------
@dp.callback_query(F.data.startswith("reply:"))
async def start_reply(cq: CallbackQuery, state: FSMContext):
    if not await is_admin(cq.from_user.id):
        return
    target_id = int(cq.data.split(":", 1)[1])
    await state.update_data(target_id=target_id)
    await state.set_state(ReplyStates.waiting_text)
    await cq.message.answer(
        f"Reply rejimi yoqildi.\nTarget user_id: <code>{target_id}</code>\n"
        f"Xabar yuboring. Chiqish: /done"
    )
    await cq.answer()

@dp.message(Command("done"))
async def finish_reply(msg: Message, state: FSMContext):
    if not await is_admin(msg.from_user.id):
        return
    await state.clear()
    await msg.answer("Reply rejimi yopildi.")

@dp.message(ReplyStates.waiting_text)
async def forward_reply(msg: Message, state: FSMContext):
    if not await is_admin(msg.from_user.id):
        return
    data = await state.get_data()
    target_id = data.get("target_id")
    if not target_id:
        await msg.answer("Target yo'q. /done qilib chiqib, tugma orqali qayta kir.")
        return
    try:
        await msg.copy_to(chat_id=target_id)
        await msg.answer(f"✔️ Yuborildi → <code>{target_id}</code>")
    except Exception as e:
        await msg.answer(f"❌ Yuborilmadi: {e}")
    # ✅ Faqat bir marta reply ishlasin, keyin avtomatik yopilsin
    await state.clear()

# ---------- Admin management (panel) ----------
@dp.callback_query(F.data == "admins:add")
async def admins_add_start(cq: CallbackQuery, state: FSMContext):
    if not await is_superadmin(cq.from_user.id):
        await cq.answer("Faqat superadmin uchun.", show_alert=True)
        return
    await state.set_state(AdminMgmtStates.waiting_add)
    await cq.message.answer("➕ Admin qo'shish: user_id yoki @username yuboring.\nYakunlash: /cancel")
    await cq.answer()

async def _resolve_user_token(token: str) -> Optional[int]:
    token = (token or "").strip()
    if token.isdigit():
        return int(token)
    # username orqali DBdan topamiz
    return await db.find_user_id_by_username(token)

@dp.message(AdminMgmtStates.waiting_add)
async def admins_add_do(msg: Message, state: FSMContext):
    if not await is_superadmin(msg.from_user.id):
        return
    uid = await _resolve_user_token(msg.text)
    if not uid:
        await msg.answer("❌ Topilmadi. user_id yoki @username yubor.")
        return
    ok = await db.add_admin(uid)
    if ok:
        await msg.answer(f"✅ Admin qo'shildi: <code>{uid}</code>")
    else:
        await msg.answer("ℹ️ Qo'shilmadi (allaqachon admin yoki superadmin).")
    await state.clear()

@dp.callback_query(F.data == "admins:remove")
async def admins_remove_start(cq: CallbackQuery, state: FSMContext):
    if not await is_superadmin(cq.from_user.id):
        await cq.answer("Faqat superadmin uchun.", show_alert=True)
        return
    await state.set_state(AdminMgmtStates.waiting_remove)
    await cq.message.answer("➖ Admin o‘chirish: user_id yoki @username yuboring.\nYakunlash: /cancel")
    await cq.answer()

@dp.message(AdminMgmtStates.waiting_remove)
async def admins_remove_do(msg: Message, state: FSMContext):
    if not await is_superadmin(msg.from_user.id):
        return
    uid = await _resolve_user_token(msg.text)
    if not uid:
        await msg.answer("❌ Topilmadi. user_id yoki @username yubor.")
        return
    ok = await db.remove_admin(uid)
    if ok:
        await msg.answer(f"✅ Admin o‘chirildi: <code>{uid}</code>")
    else:
        await msg.answer("ℹ️ O‘chirilmadi (admin emas yoki superadmin).")
    await state.clear()

@dp.callback_query(F.data == "admins:list")
async def admins_list(cq: CallbackQuery):
    if not await is_superadmin(cq.from_user.id):
        await cq.answer("Faqat superadmin uchun.", show_alert=True)
        return
    txt = await db.list_admins_text()
    await cq.message.answer(txt)
    await cq.answer()

# ---------- Quick actions (promote/revoke) ----------
@dp.callback_query(F.data.startswith("admins:promote:"))
async def quick_promote(cq: CallbackQuery):
    if not await is_superadmin(cq.from_user.id):
        await cq.answer("Faqat superadmin uchun.", show_alert=True)
        return
    uid = int(cq.data.split(":")[2])
    ok = await db.add_admin(uid)
    if ok:
        await cq.answer("Qo'shildi ✅", show_alert=False)
        await cq.message.answer(f"✅ Admin qo'shildi: <code>{uid}</code>")
    else:
        await cq.answer("Qo'shilmadi (allaqachon admin yoki superadmin).", show_alert=True)

@dp.callback_query(F.data.startswith("admins:revoke:"))
async def quick_revoke(cq: CallbackQuery):
    if not await is_superadmin(cq.from_user.id):
        await cq.answer("Faqat superadmin uchun.", show_alert=True)
        return
    uid = int(cq.data.split(":")[2])
    ok = await db.remove_admin(uid)
    if ok:
        await cq.answer("O‘chirildi ✅", show_alert=False)
        await cq.message.answer(f"✅ Admin o‘chirildi: <code>{uid}</code>")
    else:
        await cq.answer("O‘chirilmadi (admin emas yoki superadmin).", show_alert=True)

# ---------- User → Admin relay ----------
@dp.message(CommandStart())
async def start_cmd(msg: Message):
    await db.upsert_user(msg.from_user)
    await msg.answer("Savolingizni yozing. Admin ko‘radi va javob beradi.")

@dp.message(F.text | F.photo | F.video | F.audio | F.document | F.sticker | F.voice | F.video_note | F.animation)
async def relay_to_admins(msg: Message):
    if await is_admin(msg.from_user.id):
        return


    await db.upsert_user(msg.from_user)

    u = msg.from_user
    username = f"@{u.username}" if u.username else "-"
    first = u.first_name or "-"
    last = u.last_name or "-"
    meta = (
        f"<b>Yangi xabar</b>\n"
        f"ID: <code>{u.id}</code>\n"
        f"Username: {username}\n"
        f"Ism: {first}\n"
        f"Familiya: {last}\n"
        f"Vaqt: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )

    admin_ids = await all_admin_ids()
    for aid in admin_ids:
        try:
            await bot.send_message(
                aid,
                meta,
                reply_markup=reply_kb(u.id, superadmin=await is_superadmin(aid)),
            )
            await msg.copy_to(chat_id=aid)
        except Exception:
            # Admin bloklagan bo'lishi mumkin — baribir davom etamiz
            pass

# ---------- Run ----------
async def main():
    print("Bot ishga tushyapti (DB mode)...")
    await db.connect()
    await db.ensure_schema()
    await db.seed_admins(SEED_SUPERADMINS, SEED_ADMINS)
    await dp.start_polling(bot)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        print("To'xtadi.")
